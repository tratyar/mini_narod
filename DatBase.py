from openpyxl import load_workbook
import sqlite3


data = {'💡 Информация': 'info', '🎞️ История': "histori", '👺 Традиции': 'traditions', '🥞 Фирменные блюда': "kooking"}


class Db_work():
    def __init__(self):
        self.wb = load_workbook('арктика.xlsx')
        self.sheet = self.wb['Лист1']

    def ret_inf(self, col, row):
        return self.sheet.cell(row=row, column=col).value

    def ret_peopl(self, cnt):
        with sqlite3.connect("Northern_peoples.sqlite") as db:
            cur = db.cursor()
            return list(map(lambda x: x[0],
                            cur.execute(f'SELECT name FROM Piopls WHERE region LIKE "%{cnt}%"').fetchall()))

    def ret_cherta(self, cnt, pepe):
        with sqlite3.connect("Northern_peoples.sqlite") as db:
            cur = db.cursor()
            return cur.execute(f'SELECT {data[cnt]} FROM Piopls WHERE id = {pepe}').fetchall()[0]